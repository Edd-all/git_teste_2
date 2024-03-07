import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font # Change the color and form of cells
import pandas as pd
import datetime


worksheets = openpyxl.load_workbook('./Draft.xlsm') # load data to the variable worksheets
script = worksheets['Script'] # acess the sheet script
drafts = worksheets['Drafts'] # acess the sheet drafts


# create a datetime object for the current date and time
today = datetime.datetime.now()
# create a datetime object for the yesterday date and time
two_days_ago = datetime.datetime.now() - datetime.timedelta(days=2)


# Fill parameters with blue color
my_fill = PatternFill(start_color='7FFF00',end_color='7FFF00',fill_type='solid')


rows_to_delete = []
for row in script.iter_rows(min_row=script.min_row, max_row=script.max_row):
    date_cell = row[4].value  # date is in the row 4, that is the "E" (índice 4)
    if date_cell is not None and isinstance(date_cell, datetime.datetime): #The isinstance() function returns True if the specified object is of the specified type, otherwise False.
        if date_cell < two_days_ago:
            rows_to_delete.append(row) #The append() method appends an element to the end of the list.
# delete rows with dates before 2 days ago
for row in reversed(rows_to_delete):  # remove rows in reverse order to avoid problems
    script.delete_rows(row[4].row)  # date is in the row 4, that is the "E" (índice 4)


#TODO - filtrar por MachX
for row in script.iter_rows(min_row=script.min_row, max_row=script.max_row):
    machx_cell = row[0].value  # date is in the row 4, that is the "E" (índice 4)
    if machx_cell



# Loop through rows in the script
for row in script.iter_rows(min_row=script.min_row, max_row=script.max_row):
    for cell in row: # Loop through cells in the row
        cell.fill = my_fill #paint cells       


# Salva o arquivo
worksheets.save('test.xlsx')